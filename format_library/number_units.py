import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, NamedStyle, numbers


def number_format(month_year):

    # Load the workbook
    file_path = r'C:\Users\ttarek\OneDrive - Tarion\Projects\Python\Investment Working - ' + month_year + '.xlsx'
    workbook = openpyxl.load_workbook(file_path)

    # main code

    # Create a named style for the number format
    number_format_style = NamedStyle(name="number_format_style")
    number_format_style.number_format = '#,##0.00'

    # Iterate through all the sheets in the workbook
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, float):
                    cell.style = number_format_style

    # for exceptions
    columns_to_format = ['Units']

    # Iterate through all sheets in the workbook
    for sheet in workbook.worksheets:
        # Iterate through all columns in the sheet
        for col in sheet.iter_cols(min_row=1, max_row=1):  # Only check the first row for column headers
            for cell in col:
                if cell.value in columns_to_format:
                    # Apply formatting to the entire column
                    for row in sheet.iter_rows(min_col=cell.column, max_col=cell.column, min_row=1, max_row=sheet.max_row):
                        for c in row:
                            c.number_format = '#,##0.00'

    # Save the workbook
    workbook.save(file_path)
